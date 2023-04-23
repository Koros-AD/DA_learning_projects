#: с помощью openpyxl написать код, который будет создавать какую-либо матрицу товаров
# и в последнем столбце автоматически рассчитывать итоговую стоимость для каждой строки (кол-во товара*цену за единицу)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# создание новой книги
wb = Workbook()
ws = wb.active

# добавление заголовков для таблицы
headers = ["Наименование товара", "Цена за шт", "Кол-во", "Итого"]
for i, header in enumerate(headers, start=1):
    col_letter = get_column_letter(i)
    ws[f"{col_letter}1"] = header

# добавление данных в таблицу
data = [
    ("Учебник", 10, 500),
    ("Карандаш", 5, 20),
    ("Рабочая Тетрадь", 10, 250),
    ("Ручка", 10, 50)
]

for i, row in enumerate(data, start=2):
    for j, value in enumerate(row, start=1):
        col_letter = get_column_letter(j)
        ws[f"{col_letter}{i}"] = value

    # рассчет итоговой стоимости для строки
    total_col_letter = get_column_letter(j + 1)
    formula = f"={col_letter}{i} * {get_column_letter(j - 1)}{i}"
    ws[f"{total_col_letter}{i}"] = formula

# сохранение книги в файл
wb.save("D:\Media/products.xlsx")
